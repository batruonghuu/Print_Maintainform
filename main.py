import xlwings as xl
import os
import tkinter.filedialog
import xlrd
import openpyxl
import pandas as pd
# import numpy

file_template_path = tkinter.filedialog.askopenfilename(title="Mở file template phiếu bảo dưỡng mẫu")
file_groupttb_path = tkinter.filedialog.askopenfilename(title="Mở file group TTB")
file_vlookup = tkinter.filedialog.askopenfilename(title="Mở file ánh xạ mã thiết bị")
# ask directory

df_groupttb = pd.read_excel(file_groupttb_path)
df_vlookup = pd.read_excel(file_vlookup)
# create a pandas dataframe from file excel

df_groupttb.fillna(method='ffill', inplace=True)
# unmergce mergedcell to seperate cells with same value

df_groupttb[['Mã trang thiết bị','Tên trang thiết bị']] = df_groupttb['Hệ thống/Trang thiết bị'].str.split(":",expand=True)
# seperate 1 column "text-to-column" 2 column

df_groupttb['name1'] = df_groupttb['Người thực hiện'].str.split(",",expand=True,n=2).iloc[:,:1]
# seperate 1 column by character "," by maximum n+1 element, and get only 1 first column

df_groupttb['Mã trang thiết bị'], df_groupttb['Tên trang thiết bị'] = df_groupttb['Tên trang thiết bị'],df_groupttb['Mã trang thiết bị']
# swap the column 'Tên trang thiết bị' to position before 'Mã trang thiết bị'

df_groupttb.drop(df_groupttb[df_groupttb['Tên trang thiết bị'] == '»'].index, inplace = True)
# delete row if value in column include >>

df_copy = df_groupttb.copy()
# create a copy of pandas dataframe

df_groupttb = df_groupttb[df_groupttb['Mã trang thiết bị'].notnull()]
# delete row if value is null

df_groupttb['Thời gian dự kiến'] = pd.to_datetime(df_groupttb['Thời gian dự kiến'], errors='coerce',dayfirst=True)
df_groupttb['Thời gian dự kiến'] = df_groupttb['Thời gian dự kiến'].dt.strftime('%Y-%m-%d')
# convert datetime column
# first: convert to datetime format (keep non-datetime format cell),
# second: convert to strings,

df_copy['xxx'] = df_copy['Tên trang thiết bị'].str[:9]
# create a column based on number of character of other column

# df_copy['Nhà ga'] = df_copy['xxx'].str.contains('string_to_check')
# create a column contains string

df_copy['Nhà ga'] = numpy.where(df_copy['xxx'].str.contains('-T1'), 'T1', 'T2')

df_copy.drop_duplicates(subset=['Thời gian dự kiến','Nhà ga'], keep='first', inplace=True)
# remove the rows with same value in 'subset' column, only keep the first row



df_copy['Thời gian dự kiến'] = pd.to_datetime(df_copy['Thời gian dự kiến'], errors='coerce',dayfirst=True)
df_copy['Thời gian dự kiến'] = df_copy['Thời gian dự kiến'].dt.strftime('%Y-%m-%d')
df_copy = df_copy[df_copy['Thời gian dự kiến'].notnull()]

df_copy = df_copy.merge(df_vlookup,on='Tên trang thiết bị', how='left')
# insert row from df_vloopup with matching based on column

set_datetime = set(df_groupttb['Thời gian dự kiến'].unique())
list_datetime = list(set_datetime)
# print(list_datetime)

df_groupttb.to_excel('dataframe.xlsx', index=False)
df_copy.to_excel('dataframe_copy.xlsx', index=False)
# Save to a excel
def insert_new_row(brow,sheet,value1,value2):
    # sheet.range(brow,1).api.EntireRow.Insert()
    # insert a new row with keeping format of previous row

    sheet.cell(row=brow,column=2).value = value1
    sheet.cell(row=brow,column=3).value = value2

    sheet.row_dimensions[brow].height = None
    # sheet.sheet_format.autofit_row_height()

    # sheet.range(brow,1).autofit()
    ## autofit row based on content

template_wb = openpyxl.load_workbook(file_template_path)
originalsheet = template_wb['Original Sheet']

for i in set(df_copy['Thời gian dự kiến']):
    for j in set(df_copy['Nhà ga']):
        check_boolean = (df_copy['Thời gian dự kiến'] == i) & (df_copy['Nhà ga'] == j)
        result1 = list(df_copy.loc[check_boolean,'Tên trang thiết bị'])
        result2 = list(df_copy.loc[check_boolean, 'Mã trang thiết bị new'])
        result3 = list(df_copy.loc[check_boolean, 'name1'])
        # return the value of first column based on determine value from other column
        if len(list(result1)) > 0:
            print(i,j,result1[0],result2[0],result3[0])
            duplicate_ws = template_wb.copy_worksheet(originalsheet)
            duplicate_ws.title = i + " " + j
            # duplicate sheet and give a name

            template_wb.active = duplicate_ws
            # active the sheet

            insert_new_row(8,duplicate_ws,result1[0],result2[0])
            row_index = 0
            # col_index = 0
            # print(len(duplicate_ws.iter_rows(values_only=True)))
            for row in duplicate_ws.iter_rows(values_only=True):
                # print(row)
                for col_index in range(len(row)):
                    if "Họ tên:…" in str(row[col_index]):
                        duplicate_ws.cell(row=row_index+1, column=col_index+1, value="Họ tên: "+str(result3[0]))
                    if "T1,T2" in str(row[col_index]):
                        if 'T2' in j:
                            duplicate_ws.cell(row=row_index+1, column=col_index+1, value="PHIẾU BẢO DƯỠNG MÁY TRẠM LÀM THỦ TỤC HÀNH KHÁCH T2")
                        else:
                            duplicate_ws.cell(row=row_index+1, column=col_index+1, value="PHIẾU BẢO DƯỠNG MÁY TRẠM LÀM THỦ TỤC HÀNH KHÁCH T1")
                    if 'Ngày thực hiện' in str(row[col_index]):
                        duplicate_ws.cell(row=row_index+1, column=col_index+1, value="Ngày thực hiện: "+str(i))
                    if 'Ngày kiểm tra' in str(row[col_index]):
                        duplicate_ws.cell(row=row_index+1, column=col_index+1, value="Ngày kiểm tra: "+str(i))

                row_index = row_index +1

template_wb.save('fileprint.xlsx')







