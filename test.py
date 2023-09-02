import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, PatternFill
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import smtplib

# read the excel file and sheet

import tkinter.filedialog
# import win32com.client as win32
#
# excel = win32.gencache.EnsureDispatch('Excel.Application')

file_template_path = tkinter.filedialog.askopenfilename(title="Mở file để in")
wb = openpyxl.load_workbook(file_template_path)
ws = wb['Original Sheet']

# extract the table data and formatting
table_data = []
table_styles = {}
for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
    table_data.append(list(row))
    # print(row)
    for cell in row:
        cell_coord = cell.coordinate
        cell_font = cell.font.copy()
        # print(cell_font)
        cell_border = cell.border.copy()
        cell_fill = cell.fill.copy()
        cell_number_format = cell.number_format
        if cell_font or cell_border or cell_fill or cell_number_format:
            table_styles[cell_coord] = {
                'font': cell_font,
                'border': cell_border,
                'fill': cell_fill,
                'number_format': cell_number_format
            }

# create an html table from the data and styles
html_table = '<table>'
for row in table_data:
    html_table += '<tr>'
    for cell in row:
        html_table += '<td>{}</td>'.format(cell)
    html_table += '</tr>'
html_table += '</table>'

# add the table styles to the html
html_styles = '<style>'
for cell_coord, styles in table_styles.items():
    font = 'font-family: {}; font-size: {}pt; '.format(styles['font'].name, styles['font'].size)
    if styles['font'].bold:
        font += 'font-weight: bold; '
    if styles['font'].italic:
        font += 'font-style: italic; '
    # border = ''
    # if styles['border'].left.style != 'none':
    #     border += 'border-left: {} {}; '.format(styles['border'].left.style, styles['border'].left.color.rgb)
    # if styles['border'].right.style != 'none':
    #     border += 'border-right: {} {}; '.format(styles['border'].right.style, styles['border'].right.color.rgb)
    # if styles['border'].top.style != 'none':
    #     border += 'border-top: {} {}; '.format(styles['border'].top.style, styles['border'].top.color.rgb)
    # if styles['border'].bottom.style != 'none':
    #     border += 'border-bottom: {} {}; '.format(styles['border'].bottom.style, styles['border'].bottom.color.rgb)
    fill = 'background-color: {}; '.format(styles['fill'].start_color.rgb)
    html_styles += 'td[style*="{}"]{{{}{}{}}}'.format(cell_coord, font, Border, fill)
html_styles += '</style>'
html_table = html_styles + html_table
html_table = '<table>' + html_table + '</table>'
print(html_table)
with open('table.html','w',encoding='utf-8') as file:
    file.write(html_table)